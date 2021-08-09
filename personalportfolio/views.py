import os
from django.shortcuts import render, redirect
from faker import Faker
import psycopg2
from openpyxl import load_workbook
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm

from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .decorators import unauthenticated_user, allowed_users
from django.contrib.auth.models import Group

@login_required(login_url='/registration/login/')
def login_request(request):
    if request.method == 'POST':
        form = AuthenticationForm(request=request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                messages.info(request, f"You are now logged in as {username}")
            return render(request, "index.html")

    form = AuthenticationForm()
    return render(request = request,
                    template_name = "login.html",
                    context={"form":form})

@unauthenticated_user
def login(request):
    """
    Creates login view
    Returns: rendered login page
    """
    return render(request, 'login.html')


def register(response):
    if response.method == 'POST':
        form = UserCreationForm(response.POST)
        if form.is_valid():
            user = form.save()           
            if response.POST.get('group') == 'current':
                group = Group.objects.get(name='only user')
            user.groups.add(group)
            return redirect("login")
    else:
        form = UserCreationForm()
    return render(response, 'registration/register.html', {'form': form})

def logout_request(request):
    logout(request)
    messages.info(request, "you have successfully logged out.")
    return redirect('login')

def home(request):
    return render(request, "index.html")


@allowed_users(allowed_roles=["creator"])
def upload(request):
    if request.method == 'POST':
        if request.FILES.get('document'):
            file = request.FILES['document']
                         
            workbook = load_workbook(filename=file, data_only=True)
            xls = workbook[workbook.sheetnames[0]] 

            connection = connect_to_db()
            sql_create_movie_table = ''' CREATE TABLE IF NOT EXISTS movie (
                                            name text PRIMARY KEY,
                                            image text,
                                            url text
                                            )'''
            fake = Faker()
            if connection is not None:
                create_table(connection, sql_create_movie_table)
                
                num_rows = xls.max_row
                while True:
                    if xls.cell(num_rows, 3).value is not None:
                        break
                    else:
                        num_rows -= 1
                for i in range(2, num_rows+1):
                    """
                        encoded_string = ''
                        with open(xls['B'+str(i)].value, 'rb') as img_f:
                        encoded_string = base64.b64encode(img_f.read())
                    """
                    student = (xls['A'+str(i)].value,
                               xls['B'+str(i)].value,
                               xls['C'+str(i)].value)
                               
                    create_movie(connection, student)
            return render(request, 'index.html')
    return render(request, 'upload.html')

def connect_to_db():
    """
    Creates a connection to the database
    Returns database connection
    """
    if 'ON_HEROKU' in os.environ:
        connection = psycopg2.connect(os.environ['DATABASE_URL'],
                                      sslmode='require')
    else:
        connection = psycopg2.connect(dbname="d6sr6kbcak5mfs",
                                      user="ojqujfrhirgypw",
                                      password="aa31f65a42841c366d1b5c9b8b763e5909be6fdf50f111bc2a48c7936f9ae18b",
                                      host="ec2-3-231-69-204.compute-1.amazonaws.com",
                                      port="5432")
    
    return connection

def create_table(conn, create_table_sql):
    """
    Creates a database table
    Inputs: database connection, SQL for table creation
    """
    try:
        cur = conn.cursor()
        cur.execute("DROP TABLE IF EXISTS movie")
        cur.execute(create_table_sql)
        conn.commit()
    except psycopg2.OperationalError as error:
        print(error)

def create_movie(conn, movie):
    """
    Inserts into movie table using SQL
    Inputs: database connection and movie tuple
    Returns: auto increment value property of row
    """
    sql = ''' INSERT INTO movie(name, image, url)
              VALUES(%s, %s, %s)'''
    cur = conn.cursor()
    cur.execute(sql, movie)
    conn.commit()
    return cur.lastrowid

@allowed_users(allowed_roles=['only user',
                              'creator'])
def movies(request):
    connection = connect_to_db()
    cursor = connection.cursor()
    sql = 'SELECT * FROM movie;'
    cursor.execute(sql)
    results = cursor.fetchall()

    results_dict = {
        'movies': results
    }

    cursor.close()
    connection.close()
    return render(request, 'movies.html', results_dict)


@allowed_users(allowed_roles=['creator'])
def store(request):
    if request.method == 'POST':
        student_id = request.POST.get('id')
        if request.POST.get('update'):
            result = (request.POST.get('name'),
                       request.POST.get('image'),
                       request.POST.get('url'))
            connection = connect_to_db()
            create_movie(connection, result)
    return render(request, 'store.html')

def PhysicalActivity(request):

    return render(request, 'physical_activity.html')

def project(request):

    return render(request, "project.html")


def question(request):
    return render(request, "question.html")




# physical activites

def walk(request):

    return render(request, 'physicalactivity/walk.html')

def jogging(request):

    return render(request, "'physicalactivity/jogging.html'")


def run(request):
    return render(request, 'physicalactivity/run.html')


# gym folder from physical activites folder
def app(request):
    return render(request, 'physicalactivity/gym/app.html')

def chess(request):
    return render(request, 'physicalactivity/gym/chess.html')

def leg(request):
    return render(request, 'physicalactivity/gym/leg.html')

def muscle(request):
    return render(request, 'physicalactivity/gym/muscle.html')



def chat(request):
    return render(request, 'chat.html')


def programming(request):
    return render(request, 'programming.html')

def python(request):
    return render(request, 'python/python.html')

def pythonsyntax(request):
    return render(request, 'python/syntax.html')

def pythoncomment(request):
    return render(request, 'python/comment.html')

def pythonvariable(request):
    return render(request, 'python/variable.html')