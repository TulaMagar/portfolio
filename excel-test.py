from openpyxl import load_workbook
wb = load_workbook(filename = 'movies-data.xlsx')

import pandas as pd
df = pd.DataFrame(file, columns=['name', 'image', 'url'])
xls = wb[wb.sheetnames[0]]
for i in xls:
    print(i[0].value)
